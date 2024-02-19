from openpyxl.styles import PatternFill

from src.excel_for_friends.movies import Movie
from src.excel_for_friends.tv_shows import Show
from src.excel_for_friends.games import Game
from src.excel_for_friends.songs import Song
from src.excel_for_friends.sort_excel import SortExcel

import openpyxl
import os

RED_FILL_COLOR = PatternFill(start_color='FFFF0000',
                             end_color='FFFF0000',
                             fill_type='solid')


class ExcelConfig(Movie, Show, Game, Song, SortExcel):
    def __init__(self, file_path, first_entry, second_entry, third_entry):
        Movie.__init__(self, first_entry, second_entry, third_entry,  RED_FILL_COLOR)
        Show.__init__(self, first_entry, second_entry, third_entry,  RED_FILL_COLOR)
        Game.__init__(self, first_entry, second_entry, third_entry,  RED_FILL_COLOR)
        Song.__init__(self, first_entry, second_entry, third_entry,  RED_FILL_COLOR)
        SortExcel.__init__(self)
        self.file_path = file_path
        self.wb = openpyxl.Workbook()

    def _open_excel_file(self):
        self.wb = self.wb

    def _load_excel_file(self):
        self.wb = openpyxl.load_workbook(self.file_path)

    def write_to_excel(self):
        Movie.add_column_names(self)
        Show.add_column_names(self)
        Game.add_column_names(self)
        Song.add_column_names(self)
        # if os.path.exists(self.file_path):
        #     # self._load_excel_file()
        #     Movie.add_column_names(self)
        #     Show.add_column_names(self)
        #     Game.add_column_names(self)
        #     Song.add_column_names(self)
        # else:
        #     # self._open_excel_file()
        #     Movie.add_column_names(self)
        #     Show.add_column_names(self)
        #     Game.add_column_names(self)
        #     Song.add_column_names(self)

        # self.choose_option()
    def choose_option(self):
        # option = self._get_option_number()
        # while int != type(option) or option < 1 or option > 5:
        #     if int == type(option):
        #         print(f"{self.warning_color}The number have to be between 1 and 4!{self.end_color}")
        #     option = self._get_option_number()

        if self.get_radio_button() == 1:
            Movie.add_values_to_cells(self)
            self.save_excel_file()
        if self.get_radio_button() == 2:
            Show.add_values_to_cells(self)
            self.save_excel_file()
        if self.get_radio_button() == 3:
            Game.add_values_to_cells(self)
            self.save_excel_file()
        if self.get_radio_button() == 4:
            Song.add_values_to_cells(self)
            self.save_excel_file()
            # answer = self._get_answer()
        SortExcel.sort_value(self, self.get_radio_button())

    # def _get_option_number(self):
    #     try:
    #         option = int(input(f"{self.success_color}Choose what do you want to add 1 - movie, 2 - tv show, 3 - game, 4 - song: {self.end_color}"))
    #         return option
    #     except ValueError:
    #         print(f"{self.warning_color}You have to type a number!{self.end_color}")

    # def _get_answer(self):
    #     try:
    #         answer = input(f"{self.success_color}Do you want to add another row?(y/n): {self.end_color}")
    #         return answer.lower()
    #     except ValueError:
    #         print(f"{self.warning_color}You have to type a number!{self.end_color}")

    def delete_sheet(self):
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

    def save_excel_file(self):
        self.delete_sheet()
        self.wb.save(f"{self.file_path}")
