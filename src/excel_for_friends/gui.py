import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import os
import openpyxl
from src.excel_for_friends.excel_config import ExcelConfig
from src.excel_for_friends.exceptions import NumberNotInRange

LABEL_FONT = ("arial", 12, "bold")
RADIO_BUTTON_FONT = ("arial", 12, "bold")
CHECK_BOX_FONT = ("arial", 8, "bold")
ENTRY_FONT = ("arial", 8)
BUTTON_FONT = ("arial", 10, "bold")
BUTTON_COLOR = "#90ee90"
FRAME_BG_COLOR = "#f0f0f0"
number = 0


class App(tk.Tk, ExcelConfig):
    def __init__(self, warning_color, information_color, success_color, end_color):
        tk.Tk.__init__(self)
        self.file_path = None
        self.first_entry = None
        self.second_entry = None
        self.third_entry = None
        ExcelConfig.__init__(self, first_entry=self.first_entry, second_entry=self.second_entry, third_entry=self.third_entry, file_path=self.file_path)
        self.title("Search")
        self.style = ttk.Style()
        self.style.configure("TButton", font=BUTTON_FONT, background=BUTTON_COLOR)

        self.first_label = None
        self.second_label = None
        self.third_label = None
        self.eText = tk.StringVar()
        self.radio_state = tk.IntVar(value=1)
        self.checked_state = tk.IntVar()
        self.create_widgets()


    # style = ttk.Style()
    # style.configure("TButton", font=BUTTON_FONT, background=BUTTON_COLOR)
    def create_widgets(self):
        self.first_row()
        self.second_row()
        self.third_row()
        self.fourth_row()
        self.fifth_row()
        self.sixth_row()

    def clear_entries(self):
        self.first_entry.delete(0, 'end')
        self.second_entry.delete(0, 'end')
        self.third_entry.delete(0, 'end')

    def load_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=(
                ("Excel Files", "*.xlsx"),
                ("Python Files", ("*.zip", "*.zip")),
                ("All Files", "*.*")
            ),
            title="Choose a file"
        )
        self.file_path = file_path
        end_name = os.path.basename(file_path)
        self.eText.set(end_name)
        self._load_excel_file()

    def create_file(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Create a file")
        self.file_path = file_path
        self._open_excel_file()
        self.wb.save(file_path)
        self.write_to_excel()
        # wb = openpyxl.Workbook()
        # wb.save(f"{file_path}")
        end_name = os.path.basename(file_path)
        self.eText.set(end_name)

    def add_to_excel_button(self):
        # self.write_to_excel()
        try:
            self.choose_option()
            if self.checkbutton_used() == 0:
                self.quit()
            else:
                messagebox.showinfo("Status", "Progress was saved!")
            self.clear_entries()
        except ValueError:
            messagebox.showerror("Error", "The rating have to be in number format!")
        except KeyError:
            messagebox.showerror("Error", "You have to choose the file!")
        except NumberNotInRange:
            messagebox.showerror("Error", "Rating have to be in range 0 to 100!")

    def get_radio_button(self):
        return self.radio_state.get()

    def radio_used(self):
        genre = ["Genre", "Genre", "Category", "Singer"]
        self.first_label['text'] = "Name:"
        if self.radio_state.get() == 1 or self.radio_state.get() == 2:
            self.second_label['text'] = f"{genre[self.radio_state.get() - 1]}:"
            self.second_label.pack(padx=12)

        elif self.radio_state.get() == 3:
            self.second_label['text'] = f"{genre[self.radio_state.get() - 1]}:"
            self.second_label.pack(padx=0)
        elif self.radio_state.get() == 4:
            self.second_label['text'] = f"{genre[self.radio_state.get() - 1]}:"
            self.second_label.pack(padx=10)

        self.third_label['text'] = "Rating:"

    def checkbutton_used(self):
        return self.checked_state.get()

    def first_row(self):
        frame1 = tk.Frame(self, bg=FRAME_BG_COLOR)
        frame1.grid(row=0, column=0, padx=10, pady=10)

        artist_button = ttk.Button(frame1, text="Open a File", command=self.load_file)
        artist_button.pack(side="left", padx=5, pady=5)
        create_button = ttk.Button(frame1, text="Create a File", command=self.create_file)
        create_button.pack(side="left", padx=5, pady=5)
        artist_entry = tk.Entry(frame1, font=LABEL_FONT, state="readonly", textvariable=self.eText, width=55)
        artist_entry.pack(side="left", padx=5, pady=5)

    def second_row(self):
        frame2 = tk.Frame(self, bg=FRAME_BG_COLOR)
        frame2.grid(row=1, column=0, padx=10, pady=10)

        movies_radio_button = tk.Radiobutton(frame2, text="Movies", value=1, variable=self.radio_state, command=self.radio_used, font=RADIO_BUTTON_FONT, indicatoron=False)
        movies_radio_button.pack(side="left", padx=40, pady=5)
        shows_radio_button = tk.Radiobutton(frame2, text="TV Shows", value=2, variable=self.radio_state, command=self.radio_used, font=RADIO_BUTTON_FONT, indicatoron=False)
        shows_radio_button.pack(side="left", padx=40, pady=5)
        games_radio_button = tk.Radiobutton(frame2, text="Games", value=3, variable=self.radio_state, command=self.radio_used, font=RADIO_BUTTON_FONT, indicatoron=False)
        games_radio_button.pack(side="left", padx=40, pady=5)
        songs_radio_button = tk.Radiobutton(frame2, text="Songs", value=4, variable=self.radio_state, command=self.radio_used, font=RADIO_BUTTON_FONT, indicatoron=False)
        songs_radio_button.pack(side="left", padx=40, pady=5)

    def third_row(self):
        frame3 = tk.Frame(self, bg=FRAME_BG_COLOR)
        frame3.grid(row=2, column=0, padx=10, pady=10)
        self.first_label = tk.Label(frame3, text="Name:", font=LABEL_FONT)
        self.first_label.pack(side="left", padx=13)
        self.first_entry = tk.Entry(frame3, width=100)
        self.first_entry.pack(side="left", padx=5, pady=5)

    def fourth_row(self):
        frame4 = tk.Frame(self, bg=FRAME_BG_COLOR)
        frame4.grid(row=3, column=0, padx=10, pady=10)
        self.second_label = tk.Label(frame4, text="Genre:", font=LABEL_FONT)
        self.second_label.pack(side="left", padx=12)
        self.second_entry = tk.Entry(frame4, width=100)
        self.second_entry.pack(side="left", padx=5, pady=5)

    def fifth_row(self):
        frame5 = tk.Frame(self, bg=FRAME_BG_COLOR)
        frame5.grid(row=4, column=0, padx=10, pady=10)
        self.third_label = tk.Label(frame5, text="Rating:", font=LABEL_FONT)
        self.third_label.pack(side="left", padx=10)
        self.third_entry = tk.Entry(frame5, width=100)
        self.third_entry.pack(side="left", padx=5, pady=5)

    def sixth_row(self):
        frame6 = tk.Frame(self, bg=FRAME_BG_COLOR)
        frame6.grid(row=5, column=0, padx=10, pady=10)
        add_button = ttk.Button(frame6, text="Add to the excel", command=self.add_to_excel_button)
        add_button.pack(side="left", padx=5, pady=5)

        # Zaškrtávací políčko

        checkbutton = tk.Checkbutton(frame6, text="add another row?", variable=self.checked_state, font=CHECK_BOX_FONT)
        checkbutton.pack(side="left", padx=5, pady=5)

# file_path="../../data/output/", file_name="list_of_hits.xlsx"

# Vytvoření rámců pro organizaci widgetů
# frame1 = tk.Frame(main_window, bg=FRAME_BG_COLOR)
# frame1.grid(row=0, column=0, padx=10, pady=10)

# frame2 = tk.Frame(main_window, bg=FRAME_BG_COLOR)
# frame2.grid(row=1, column=0, padx=10, pady=10)

# frame3 = tk.Frame(main_window, bg=FRAME_BG_COLOR)
# frame3.grid(row=2, column=0, padx=10, pady=10)

# frame4 = tk.Frame(main_window, bg=FRAME_BG_COLOR)
# frame4.grid(row=3, column=0, padx=10, pady=10)

# frame5 = tk.Frame(main_window, bg=FRAME_BG_COLOR)
# frame5.grid(row=4, column=0, padx=10, pady=10)

# frame6 = tk.Frame(main_window, bg=FRAME_BG_COLOR)
# frame6.grid(row=5, column=0, padx=10, pady=10)

# Entry pro cestu k souboru
# eText = tk.StringVar()
#
# artist_button = ttk.Button(frame1, text="Open a File", command=load_file)
# artist_button.pack(side="left", padx=5, pady=5)
# create_button = ttk.Button(frame1, text="Create a File", command=create_file)
# create_button.pack(side="left", padx=5, pady=5)
# artist_entry = tk.Entry(frame1, font=LABEL_FONT, state="readonly", textvariable=eText, width=55)
# artist_entry.pack(side="left", padx=5, pady=5)

# Rádiová tlačítka
# radio_state = tk.IntVar(value=1)
# movies_radio_button = tk.Radiobutton(frame2, text="Movies", value=1, variable=radio_state, command=radio_used, font=RADIO_BUTTON_FONT, indicatoron=False)
# movies_radio_button.pack(side="left", padx=40, pady=5)
# shows_radio_button = tk.Radiobutton(frame2, text="TV Shows", value=2, variable=radio_state, command=radio_used, font=RADIO_BUTTON_FONT, indicatoron=False)
# shows_radio_button.pack(side="left", padx=40, pady=5)
# games_radio_button = tk.Radiobutton(frame2, text="Games", value=3, variable=radio_state, command=radio_used, font=RADIO_BUTTON_FONT, indicatoron=False)
# games_radio_button.pack(side="left", padx=40, pady=5)
# songs_radio_button = tk.Radiobutton(frame2, text="Songs", value=4, variable=radio_state, command=radio_used, font=RADIO_BUTTON_FONT, indicatoron=False)
# songs_radio_button.pack(side="left", padx=40, pady=5)

# První štítek a vstup
# first_label = tk.Label(frame3, text="Name:", font=LABEL_FONT)
# first_label.pack(side="left", padx=13)
# first_entry = tk.Entry(frame3, width=100)
# first_entry.pack(side="left", padx=5, pady=5)

# Druhý štítek a vstup
# second_label = tk.Label(frame4, text="Genre:", font=LABEL_FONT)
# second_label.pack(side="left", padx=12)
# second_entry = tk.Entry(frame4, width=100)
# second_entry.pack(side="left", padx=5, pady=5)

# Druhý štítek a vstup
# third_label = tk.Label(frame5, text="Rating:", font=LABEL_FONT)
# third_label.pack(side="left", padx=10)
# third_entry = tk.Entry(frame5, width=100)
# third_entry.pack(side="left", padx=5, pady=5)

# add_button = ttk.Button(frame6, text="Add to the excel", command=add_to_excel)
# add_button.pack(side="left", padx=5, pady=5)
#
# # Zaškrtávací políčko
# checked_state = tk.IntVar()
# checkbutton = tk.Checkbutton(frame6, text="add another row?", variable=checked_state, font=CHECK_BOX_FONT)
# checkbutton.pack(side="left", padx=5, pady=5)

# main_window.mainloop()

# import tkinter as tk
# from tkinter import filedialog
# import os
# import openpyxl
#
# LABEL_FONT = ("arial", 12, "bold")
# ENTRY_FONT = ("arial", 8)
#
# main_window = tk.Tk()
#
# main_window.title("Search")
# eText = tk.StringVar()
#
#
# # main_window.geometry("600x500")
#
#
# # artist_label = tk.Label(main_window, text="Enter the name of the artist :", font=LABEL_FONT)
# # artist_label.grid(row=1, column=0, sticky="W")
# def load_file():
#     file_path = filedialog.askopenfilename(
#         filetypes=(
#             ("Excel Files", "*.xlsx"),
#             ("Python Files", ("*.zip", "*.zip")),
#             ("All Files", "*.*")
#         ),
#         title="Choose an file"
#     )
#     end_name = os.path.basename(file_path)
#     eText.set(end_name)
#
#
# def create_file():
#     file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Create a file")
#     wb = openpyxl.Workbook()
#     wb.save(f"{file_path}")
#     end_name = os.path.basename(file_path)
#
#     # artist_entry['text'] = f"Path: {file_path}"
#     eText.set(end_name)
#
#
# artist_entry = tk.Entry(main_window, font=LABEL_FONT, state="readonly", textvariable=eText)  # width=50
# artist_entry.grid(row=0, column=0)
# artist_button = tk.Button(main_window, text="Open a File", command=load_file)
# artist_button.grid(row=0, column=1)
# create = tk.Button(main_window, text="Create a File", command=create_file)
# create.grid(row=0, column=2)
#
#
# def radio_used():
#     print(radio_state.get())
#
#
# radio_state = tk.IntVar()
# movies_radio_button = tk.Radiobutton(main_window, text="Movies", value=1, variable=radio_state, command=radio_used)
# movies_radio_button.grid(row=1, column=0)
# shows_radio_button = tk.Radiobutton(main_window, text="TV Shows", value=2, variable=radio_state, command=radio_used)
# shows_radio_button.grid(row=1, column=0, sticky="E")
# games_radio_button = tk.Radiobutton(main_window, text="Games", value=3, variable=radio_state, command=radio_used)
# games_radio_button.grid(row=1, column=1, columnspan=1)
# songs_radio_button = tk.Radiobutton(main_window, text="Songs", value=4, variable=radio_state, command=radio_used)
# songs_radio_button.grid(row=1, column=2, sticky="W")
#
#
# first_label = tk.Label(text="First Label", font=LABEL_FONT)
# first_entry = tk.Text(height=3, width=30)
#
# first_label.grid(row=2, column=0)
# first_entry.grid(row=2, column=1, columnspan=3)
#
#
# second_label = tk.Label(text="Second Label", font=LABEL_FONT)
# second_entry = tk.Entry(main_window, font=LABEL_FONT)
#
# second_label.grid(row=3, column=0)
# second_entry.grid(row=3, column=1, pady=10)
#
#
#
# def checkbutton_used():
#     # Prints 1 if On button checked, otherwise 0.
#     print(checked_state.get())
#     if checked_state.get() == 0:
#         main_window.quit()
#
#
# # variable to hold on to checked state, 0 is off, 1 is on.
# checked_state = tk.IntVar()
# checkbutton = tk.Checkbutton(text="Is On?", variable=checked_state, command=checkbutton_used)
# checked_state.get()
# checkbutton.grid(row=5, column=0)
#
# main_window.mainloop()


# import tkinter as tk
# from tkinter.filedialog import askopenfilename
#
# filename = None
#
#
# def UploadAction(event=None):
#     filename = askopenfilename()
#     print('Selected:', filename)
#     # Change text of label
#     label1['text'] = filename
#
#
# root = tk.Tk()
#
# button1 = tk.Button(text='Click Me', command=UploadAction, bg='brown', fg='white')
# button1.pack(padx=2, pady=5)
# label1 = tk.Label(text='Please choose a file')
# label1.pack(padx=2, pady=2)
#
# root.mainloop()


# import tkinter as tk
# from tkinter import filedialog
# import os
# import openpyxl
#
# LABEL_FONT = ("arial", 15, "bold")
# ENTRY_FONT = ("arial", 15)
#
# main_window = tk.Tk()
#
# main_window.title("Search")
# eText = tk.StringVar()
#
#
# def load_file():
#     file_path = filedialog.askopenfilename(
#         filetypes=(
#             ("Excel Files", "*.xlsx"),
#             ("Python Files", ("*.zip", "*.zip")),
#             ("All Files", "*.*")
#         ),
#         title="Choose a file"
#     )
#     end_name = os.path.basename(file_path)
#     eText.set(end_name)
#
#
# def create_file():
#     file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Create a file")
#     wb = openpyxl.Workbook()
#     wb.save(f"{file_path}")
#     end_name = os.path.basename(file_path)
#     eText.set(end_name)
#
#
# artist_entry = tk.Entry(main_window, font=LABEL_FONT, state="readonly", textvariable=eText)
# artist_entry.pack(side="left")
# artist_button = tk.Button(main_window, text="Open a File", command=load_file)
# artist_button.pack(side="left")
# create = tk.Button(main_window, text="Create a File", command=create_file)
# create.pack(side="left")
#
#
# def radio_used():
#     print(radio_state.get())
#
#
# radio_state = tk.IntVar()
# movies_radio_button = tk.Radiobutton(main_window, text="Movies", value=1, variable=radio_state, command=radio_used)
# movies_radio_button.pack(side="left")
# shows_radio_button = tk.Radiobutton(main_window, text="TV Shows", value=2, variable=radio_state, command=radio_used)
# shows_radio_button.pack(side="left")
# games_radio_button = tk.Radiobutton(main_window, text="Games", value=3, variable=radio_state, command=radio_used)
# games_radio_button.pack(side="left")
# songs_radio_button = tk.Radiobutton(main_window, text="Songs", value=4, variable=radio_state, command=radio_used)
# songs_radio_button.pack(side="left")
#
#
# first_label = tk.Label(text="First Label", font=LABEL_FONT)
# first_entry = tk.Text(height=3, width=30)
#
# first_label.pack(side="left")
# first_entry.pack(side="left")
#
#
# def checkbutton_used():
#     print(checked_state.get())
#     if checked_state.get() == 0:
#         main_window.quit()
#
#
# checked_state = tk.IntVar()
# checkbutton = tk.Checkbutton(text="Is On?", variable=checked_state, command=checkbutton_used)
# checked_state.get()
# checkbutton.pack(side="left")
#
# main_window.mainloop()
